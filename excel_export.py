"""
Module xuất báo cáo SWOT ra file Excel với biểu đồ và format chuyên nghiệp
"""
import pandas as pd
import plotly.graph_objects as go
import plotly.express as px
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from io import BytesIO
import base64
from typing import Dict, Any, List, Optional
import streamlit as st

# Thử import matplotlib làm fallback
try:
    import matplotlib
    matplotlib.use('Agg')  # Non-interactive backend
    import matplotlib.pyplot as plt
    HAS_MATPLOTLIB = True
except ImportError:
    HAS_MATPLOTLIB = False


def create_swot_charts(swot_data: Dict[str, Any]) -> Dict[str, BytesIO]:
    """
    Tạo các biểu đồ từ dữ liệu SWOT và lưu dưới dạng hình ảnh
    Sử dụng Plotly với fallback sang matplotlib nếu không có Chrome/Kaleido
    
    Args:
        swot_data: Dict chứa SWOT_Analysis
    
    Returns:
        Dict chứa các BytesIO object của biểu đồ (có thể rỗng nếu không thể tạo charts)
    """
    swot = swot_data.get("SWOT_Analysis", {})
    charts = {}
    
    # 1. Pie chart phân bố SWOT
    categories = ['Strengths', 'Weaknesses', 'Opportunities', 'Threats']
    counts = [
        len(swot.get("Strengths", [])),
        len(swot.get("Weaknesses", [])),
        len(swot.get("Opportunities", [])),
        len(swot.get("Threats", []))
    ]
    
    colors = ['#2ecc71', '#e74c3c', '#3498db', '#f39c12']
    
    # Thử dùng Plotly trước
    try:
        fig_pie = go.Figure(data=[go.Pie(
            labels=categories,
            values=counts,
            hole=0.4,
            marker_colors=colors,
            textinfo='label+percent+value'
        )])
        fig_pie.update_layout(
            title='Phân bố SWOT Analysis',
            width=600,
            height=400,
            showlegend=True
        )
        
        # Export pie chart
        pie_img = fig_pie.to_image(format="png", width=600, height=400)
        charts['pie_chart'] = BytesIO(pie_img)
    except (RuntimeError, Exception) as e:
        # Nếu lỗi (thường là do thiếu Chrome), thử dùng matplotlib
        if "Chrome" in str(e) or "kaleido" in str(e).lower():
            if HAS_MATPLOTLIB:
                try:
                    fig, ax = plt.subplots(figsize=(6, 4))
                    ax.pie(counts, labels=categories, colors=colors, autopct='%1.1f%%', startangle=90)
                    ax.set_title('Phân bố SWOT Analysis')
                    
                    pie_buffer = BytesIO()
                    plt.savefig(pie_buffer, format='png', dpi=100, bbox_inches='tight')
                    pie_buffer.seek(0)
                    charts['pie_chart'] = pie_buffer
                    plt.close(fig)
                except Exception as e2:
                    # Nếu matplotlib cũng lỗi, bỏ qua chart này
                    pass
        # Nếu không phải lỗi Chrome, bỏ qua chart này
    
    # 2. Bar chart Impact/Risk Level
    impact_levels = {'High': 0, 'Medium': 0, 'Low': 0}
    
    for item in swot.get("Strengths", []) + swot.get("Weaknesses", []):
        impact = item.get("impact", "Medium")
        if impact in impact_levels:
            impact_levels[impact] += 1
    
    for item in swot.get("Threats", []):
        risk = item.get("risk_level", "Medium")
        if risk in impact_levels:
            impact_levels[risk] += 1
    
    # Thử dùng Plotly trước
    try:
        fig_bar = go.Figure(data=[
            go.Bar(
                x=list(impact_levels.keys()),
                y=list(impact_levels.values()),
                marker_color=['#e74c3c', '#f39c12', '#2ecc71'],
                text=list(impact_levels.values()),
                textposition='auto'
            )
        ])
        fig_bar.update_layout(
            title='Phân bố Mức độ Ảnh hưởng/Rủi ro',
            xaxis_title='Mức độ',
            yaxis_title='Số lượng',
            width=600,
            height=400
        )
        
        # Export bar chart
        bar_img = fig_bar.to_image(format="png", width=600, height=400)
        charts['bar_chart'] = BytesIO(bar_img)
    except (RuntimeError, Exception) as e:
        # Nếu lỗi (thường là do thiếu Chrome), thử dùng matplotlib
        if "Chrome" in str(e) or "kaleido" in str(e).lower():
            if HAS_MATPLOTLIB:
                try:
                    fig, ax = plt.subplots(figsize=(6, 4))
                    bars = ax.bar(list(impact_levels.keys()), list(impact_levels.values()), 
                                  color=['#e74c3c', '#f39c12', '#2ecc71'])
                    ax.set_title('Phân bố Mức độ Ảnh hưởng/Rủi ro')
                    ax.set_xlabel('Mức độ')
                    ax.set_ylabel('Số lượng')
                    
                    # Thêm số trên mỗi cột
                    for bar in bars:
                        height = bar.get_height()
                        ax.text(bar.get_x() + bar.get_width()/2., height,
                               f'{int(height)}', ha='center', va='bottom')
                    
                    bar_buffer = BytesIO()
                    plt.savefig(bar_buffer, format='png', dpi=100, bbox_inches='tight')
                    bar_buffer.seek(0)
                    charts['bar_chart'] = bar_buffer
                    plt.close(fig)
                except Exception as e2:
                    # Nếu matplotlib cũng lỗi, bỏ qua chart này
                    pass
        # Nếu không phải lỗi Chrome, bỏ qua chart này
    
    # 3. Priority Matrix Chart (Enterprise)
    try:
        items = []
        color_map = {
            'Strengths': '#2ecc71',
            'Weaknesses': '#e74c3c', 
            'Opportunities': '#3498db',
            'Threats': '#f39c12'
        }
        impact_map = {'High': 3, 'Medium': 2, 'Low': 1}
        
        for category in ['Strengths', 'Weaknesses', 'Opportunities', 'Threats']:
            for item in swot.get(category, []):
                impact = item.get('impact') or item.get('risk_level', 'Medium')
                items.append({
                    'topic': item.get('topic', 'N/A'),
                    'category': category,
                    'priority_score': item.get('priority_score', 5),
                    'impact_score': impact_map.get(impact, 2),
                    'color': color_map[category]
                })
        
        if items:
            import numpy as np
            np.random.seed(42)
            
            fig_priority = go.Figure()
            
            for category in ['Strengths', 'Weaknesses', 'Opportunities', 'Threats']:
                cat_items = [i for i in items if i['category'] == category]
                if cat_items:
                    x_vals = [i['impact_score'] + np.random.uniform(-0.15, 0.15) for i in cat_items]
                    y_vals = [i['priority_score'] + np.random.uniform(-0.2, 0.2) for i in cat_items]
                    
                    fig_priority.add_trace(go.Scatter(
                        x=x_vals,
                        y=y_vals,
                        mode='markers',
                        marker=dict(size=14, color=color_map[category]),
                        name=category,
                        text=[i['topic'] for i in cat_items],
                        hovertemplate='<b>%{text}</b><br>Impact: %{x:.0f}<br>Priority: %{y:.1f}<extra></extra>'
                    ))
            
            fig_priority.add_hline(y=5, line_dash="dash", line_color="gray", opacity=0.5)
            fig_priority.add_vline(x=2, line_dash="dash", line_color="gray", opacity=0.5)
            
            fig_priority.update_layout(
                title='Ma trận Ưu tiên (Impact vs Priority)',
                xaxis_title='Mức độ Ảnh hưởng',
                yaxis_title='Điểm Ưu tiên',
                xaxis=dict(tickmode='array', tickvals=[1, 2, 3], ticktext=['Low', 'Medium', 'High'], range=[0.5, 3.5]),
                yaxis=dict(range=[0, 10]),
                width=700,
                height=500,
                showlegend=True,
                legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="center", x=0.5)
            )
            
            priority_img = fig_priority.to_image(format="png", width=700, height=500)
            charts['priority_matrix'] = BytesIO(priority_img)
    except Exception as e:
        pass  # Skip if chart creation fails
    
    # 4. Action Timeline Chart (Enterprise)
    try:
        action_plan = swot_data.get("Strategic_Action_Plan", [])
        if action_plan:
            actions = []
            priorities = []
            colors_list = []
            timelines = []
            
            color_map_action = {
                'Leverage Strength': '#2ecc71',
                'Address Weakness': '#e74c3c',
                'Capture Opportunity': '#3498db',
                'Mitigate Threat': '#f39c12'
            }
            
            for action in action_plan[:8]:  # Limit to 8 actions for better display
                action_text = action.get('action', 'N/A')
                # Truncate at 35 chars for Excel chart
                if len(action_text) > 35:
                    action_text = action_text[:35] + '...'
                actions.append(action_text)
                priorities.append(action.get('priority', 10))
                colors_list.append(color_map_action.get(action.get('type', ''), '#95a5a6'))
                timelines.append(action.get('timeline', 'N/A'))
            
            fig_action = go.Figure()
            fig_action.add_trace(go.Bar(
                y=actions[::-1],
                x=priorities[::-1],
                orientation='h',
                marker=dict(color=colors_list[::-1]),
                text=timelines[::-1],
                textposition='inside',
                hovertemplate='%{y}<br>Priority: %{x}<br>Timeline: %{text}<extra></extra>'
            ))
            
            fig_action.update_layout(
                title='Kế hoạch Hành động theo Ưu tiên',
                xaxis_title='Mức độ Ưu tiên',
                width=900,
                height=max(350, len(actions) * 45),
                margin=dict(l=300, r=50, t=60, b=60),
                showlegend=False
            )
            
            action_img = fig_action.to_image(format="png", width=900, height=max(350, len(actions) * 45))
            charts['action_timeline'] = BytesIO(action_img)

    except Exception as e:
        pass  # Skip if chart creation fails
    
    return charts



def export_swot_to_excel(swot_data: Dict[str, Any], df: pd.DataFrame = None, 
                         file_info: List[Dict] = None) -> BytesIO:
    """
    Xuất báo cáo SWOT ra file Excel với format chuyên nghiệp
    
    Args:
        swot_data: Dict chứa SWOT_Analysis và Executive_Summary
        df: DataFrame chứa dữ liệu đánh giá (optional)
        file_info: List thông tin các file (optional)
    
    Returns:
        BytesIO object chứa file Excel
    """
    wb = Workbook()
    wb.remove(wb.active)  # Xóa sheet mặc định
    
    # Tạo các biểu đồ (có thể rỗng nếu không có Chrome/Kaleido)
    try:
        charts = create_swot_charts(swot_data)
    except Exception as e:
        # Nếu không thể tạo charts, tiếp tục mà không có charts
        charts = {}
        if st:
            st.warning(f"⚠️ Không thể tạo biểu đồ (thiếu Chrome/Kaleido). File Excel sẽ không có charts. Lỗi: {str(e)[:100]}")
    
    # ========== SHEET 1: EXECUTIVE SUMMARY ==========
    ws_summary = wb.create_sheet("Tóm tắt Điều hành", 0)
    
    # Header
    ws_summary.merge_cells('A1:D1')
    cell = ws_summary['A1']
    cell.value = "📊 BÁO CÁO PHÂN TÍCH SWOT"
    cell.font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='1f77b4', end_color='1f77b4', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws_summary.row_dimensions[1].height = 30
    
    # Executive Summary
    ws_summary['A3'] = "📝 TÓM TẮT ĐIỀU HÀNH"
    ws_summary['A3'].font = Font(name='Arial', size=14, bold=True)
    ws_summary.row_dimensions[3].height = 25
    
    ws_summary['A4'] = swot_data.get("Executive_Summary", "Không có tóm tắt")
    ws_summary['A4'].font = Font(name='Arial', size=11)
    ws_summary['A4'].alignment = Alignment(wrap_text=True, vertical='top')
    ws_summary.merge_cells('A4:D10')
    ws_summary.row_dimensions[4].height = 100
    
    # Thống kê tổng quan
    ws_summary['A12'] = "📊 THỐNG KÊ TỔNG QUAN"
    ws_summary['A12'].font = Font(name='Arial', size=14, bold=True)
    
    swot = swot_data.get("SWOT_Analysis", {})
    stats = [
        ["Chỉ số", "Số lượng"],
        ["Strengths (Điểm mạnh)", len(swot.get("Strengths", []))],
        ["Weaknesses (Điểm yếu)", len(swot.get("Weaknesses", []))],
        ["Opportunities (Cơ hội)", len(swot.get("Opportunities", []))],
        ["Threats (Thách thức)", len(swot.get("Threats", []))],
    ]
    
    for row_idx, row_data in enumerate(stats, start=13):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws_summary.cell(row=row_idx, column=col_idx)
            cell.value = value
            if row_idx == 13:  # Header row
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
    
    # Chèn biểu đồ Pie
    if 'pie_chart' in charts:
        charts['pie_chart'].seek(0)
        img = Image(charts['pie_chart'])
        img.width = 400
        img.height = 300
        ws_summary.add_image(img, 'F3')
    
    # Chèn biểu đồ Bar
    if 'bar_chart' in charts:
        charts['bar_chart'].seek(0)
        img = Image(charts['bar_chart'])
        img.width = 400
        img.height = 300
        ws_summary.add_image(img, 'F13')
    
    # Enterprise Charts - Priority Matrix
    if 'priority_matrix' in charts:
        charts['priority_matrix'].seek(0)
        img = Image(charts['priority_matrix'])
        img.width = 500
        img.height = 350
        ws_summary.add_image(img, 'A23')
        ws_summary['A22'] = "MA TRẬN ƯU TIÊN"
        ws_summary['A22'].font = Font(name='Arial', size=14, bold=True)
    
    # Enterprise Charts - Action Timeline
    if 'action_timeline' in charts:
        charts['action_timeline'].seek(0)
        img = Image(charts['action_timeline'])
        img.width = 600
        img.height = 350
        ws_summary.add_image(img, 'F23')
        ws_summary['F22'] = "KẾ HOẠCH HÀNH ĐỘNG"
        ws_summary['F22'].font = Font(name='Arial', size=14, bold=True)
    
    # Điều chỉnh độ rộng cột
    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 15
    ws_summary.column_dimensions['C'].width = 15
    ws_summary.column_dimensions['D'].width = 15

    
    # ========== SHEET 2: KEY INSIGHTS ==========
    key_insights = swot_data.get("Key_Insights", [])
    if key_insights:
        ws_insights = wb.create_sheet("Key Insights")
        ws_insights['A1'] = "KEY INSIGHTS"
        ws_insights['A1'].font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
        ws_insights['A1'].fill = PatternFill(start_color='9b59b6', end_color='9b59b6', fill_type='solid')
        ws_insights.merge_cells('A1:B1')
        ws_insights.row_dimensions[1].height = 30
        
        for idx, insight in enumerate(key_insights, start=2):
            ws_insights.cell(row=idx, column=1).value = f"{idx-1}."
            ws_insights.cell(row=idx, column=2).value = insight
            ws_insights.cell(row=idx, column=2).alignment = Alignment(wrap_text=True)
        
        ws_insights.column_dimensions['A'].width = 5
        ws_insights.column_dimensions['B'].width = 80
    
    # ========== SHEET 3: STRENGTHS (với cột mở rộng) ==========
    ws_strengths = wb.create_sheet("Strengths")
    _create_swot_sheet(ws_strengths, "STRENGTHS (ĐIỂM MẠNH)", 
                       swot.get("Strengths", []), 
                       ['topic', 'description', 'impact', 'priority_score', 'leverage_strategy', 'kpi_metrics'])
    
    # ========== SHEET 4: WEAKNESSES (với cột mở rộng) ==========
    ws_weaknesses = wb.create_sheet("Weaknesses")
    _create_swot_sheet(ws_weaknesses, "WEAKNESSES (ĐIỂM YẾU)", 
                       swot.get("Weaknesses", []), 
                       ['topic', 'description', 'root_cause', 'impact', 'priority_score', 'mitigation_plan', 'improvement_cost'])
    
    # ========== SHEET 5: OPPORTUNITIES (với cột mở rộng) ==========
    ws_opportunities = wb.create_sheet("Opportunities")
    _create_swot_sheet(ws_opportunities, "OPPORTUNITIES (CƠ HỘI)", 
                       swot.get("Opportunities", []), 
                       ['topic', 'description', 'action_idea', 'priority_score', 'market_size', 'time_to_capture', 'required_investment'])
    
    # ========== SHEET 6: THREATS (với cột mở rộng) ==========
    ws_threats = wb.create_sheet("Threats")
    _create_swot_sheet(ws_threats, "THREATS (THÁCH THỨC)", 
                       swot.get("Threats", []), 
                       ['topic', 'description', 'risk_level', 'probability', 'severity', 'priority_score', 'contingency_plan'])
    
    # ========== SHEET 7: TOWS MATRIX (always create) ==========
    tows = swot_data.get("TOWS_Matrix", {})
    ws_tows = wb.create_sheet("TOWS Matrix")
    ws_tows['A1'] = "TOWS MATRIX - CHIẾN LƯỢC KẾT HỢP"
    ws_tows['A1'].font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
    ws_tows['A1'].fill = PatternFill(start_color='27ae60', end_color='27ae60', fill_type='solid')
    ws_tows.merge_cells('A1:C1')
    ws_tows.row_dimensions[1].height = 30
    
    row_idx = 3
    has_strategies = False
    for strategy_type, strategies in [
        ("SO Strategies (Tấn công)", tows.get("SO_Strategies", [])),
        ("WO Strategies (Chuyển đổi)", tows.get("WO_Strategies", [])),
        ("ST Strategies (Đa dạng hóa)", tows.get("ST_Strategies", [])),
        ("WT Strategies (Phòng thủ)", tows.get("WT_Strategies", []))
    ]:
        ws_tows.cell(row=row_idx, column=1).value = strategy_type
        ws_tows.cell(row=row_idx, column=1).font = Font(bold=True)
        ws_tows.cell(row=row_idx, column=1).fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
        row_idx += 1
        
        if strategies:
            has_strategies = True
        for s in strategies:
            ws_tows.cell(row=row_idx, column=1).value = s.get('strategy', '')
            ws_tows.cell(row=row_idx, column=1).alignment = Alignment(wrap_text=True)
            ws_tows.cell(row=row_idx, column=2).value = s.get('strength_used', '') or s.get('weakness_addressed', '')
            ws_tows.cell(row=row_idx, column=3).value = s.get('opportunity_leveraged', '') or s.get('threat_mitigated', '')
            row_idx += 1
        row_idx += 1
    
    if not has_strategies:
        ws_tows.cell(row=3, column=1).value = "Không có dữ liệu chiến lược TOWS"
    
    ws_tows.column_dimensions['A'].width = 50
    ws_tows.column_dimensions['B'].width = 30
    ws_tows.column_dimensions['C'].width = 30

    
    # ========== SHEET 8: ACTION PLAN (always create) ==========
    action_plan = swot_data.get("Strategic_Action_Plan", [])
    ws_action = wb.create_sheet("Action Plan")
    ws_action['A1'] = "KẾ HOẠCH HÀNH ĐỘNG CHIẾN LƯỢC"
    ws_action['A1'].font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
    ws_action['A1'].fill = PatternFill(start_color='e67e22', end_color='e67e22', fill_type='solid')
    ws_action.merge_cells('A1:G1')
    ws_action.row_dimensions[1].height = 30
    
    headers = ['Ưu tiên', 'Hành động', 'Loại', 'Timeline', 'Người phụ trách', 'Đầu tư', 'KPIs']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_action.cell(row=2, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
    
    if action_plan:
        for row_idx, action in enumerate(action_plan, start=3):
            ws_action.cell(row=row_idx, column=1).value = action.get('priority', '')
            ws_action.cell(row=row_idx, column=2).value = action.get('action', '')
            ws_action.cell(row=row_idx, column=2).alignment = Alignment(wrap_text=True)
            ws_action.cell(row=row_idx, column=3).value = action.get('type', '')
            ws_action.cell(row=row_idx, column=4).value = action.get('timeline', '')
            ws_action.cell(row=row_idx, column=5).value = action.get('owner_role', '')
            ws_action.cell(row=row_idx, column=6).value = action.get('estimated_investment', '')
            kpis = action.get('kpis', [])
            ws_action.cell(row=row_idx, column=7).value = ', '.join(kpis) if isinstance(kpis, list) else str(kpis)
            ws_action.row_dimensions[row_idx].height = 40
    else:
        ws_action.cell(row=3, column=1).value = "Không có kế hoạch hành động"
    
    ws_action.column_dimensions['A'].width = 10
    ws_action.column_dimensions['B'].width = 50
    ws_action.column_dimensions['C'].width = 20
    ws_action.column_dimensions['D'].width = 15
    ws_action.column_dimensions['E'].width = 20
    ws_action.column_dimensions['F'].width = 15
    ws_action.column_dimensions['G'].width = 40
    
    # ========== SHEET 9: COMPETITIVE ANALYSIS (always create) ==========
    competitive = swot_data.get("Competitive_Analysis", {})
    ws_comp = wb.create_sheet("Competitive")
    ws_comp['A1'] = "PHÂN TÍCH CẠNH TRANH"
    ws_comp['A1'].font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
    ws_comp['A1'].fill = PatternFill(start_color='3498db', end_color='3498db', fill_type='solid')
    ws_comp.merge_cells('A1:C1')
    ws_comp.row_dimensions[1].height = 30
    
    if competitive:
        ws_comp.cell(row=3, column=1).value = "Điểm tổng thể của bạn"
        ws_comp.cell(row=3, column=2).value = competitive.get('my_overall', 'N/A')
        ws_comp.cell(row=4, column=1).value = "Điểm đối thủ trung bình"
        ws_comp.cell(row=4, column=2).value = competitive.get('competitor_overall', 'N/A')
        ws_comp.cell(row=5, column=1).value = "Lợi thế cạnh tranh"
        ws_comp.cell(row=5, column=2).value = "Có" if competitive.get('competitive_advantage') else "Không"
        
        ws_comp.cell(row=7, column=1).value = "Tiêu chí"
        ws_comp.cell(row=7, column=1).font = Font(bold=True)
        ws_comp.cell(row=7, column=2).value = "Khoảng cách"
        ws_comp.cell(row=7, column=2).font = Font(bold=True)
        
        gaps = competitive.get('advantage_gaps', {})
        for row_idx, (key, value) in enumerate(gaps.items(), start=8):
            ws_comp.cell(row=row_idx, column=1).value = key.capitalize()
            ws_comp.cell(row=row_idx, column=2).value = value
    else:
        ws_comp.cell(row=3, column=1).value = "Không có dữ liệu cạnh tranh"
    
    ws_comp.column_dimensions['A'].width = 30
    ws_comp.column_dimensions['B'].width = 20

    
    # ========== SHEET 10: DỮ LIỆU GỐC (nếu có) ==========
    if df is not None and len(df) > 0:
        ws_data = wb.create_sheet("Dữ liệu Gốc")
        
        # Header
        ws_data['A1'] = "DỮ LIỆU ĐÁNH GIÁ GỐC"
        ws_data['A1'].font = Font(name='Arial', size=14, bold=True)
        ws_data.row_dimensions[1].height = 25
        
        # Ghi dữ liệu
        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = ws_data.cell(row=2, column=col_idx)
            cell.value = col_name
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
        
        for row_idx, row_data in enumerate(df.values, start=3):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws_data.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
        
        # Điều chỉnh độ rộng cột
        from openpyxl.utils import get_column_letter
        for col_idx in range(1, len(df.columns) + 1):
            col_letter = get_column_letter(col_idx)
            ws_data.column_dimensions[col_letter].width = 30
    
    # ========== SHEET 11: THỐNG KÊ FILE (nếu có) ==========
    if file_info:
        ws_files = wb.create_sheet("Thống kê File")
        
        # Header
        ws_files['A1'] = "THỐNG KÊ TỪNG FILE"
        ws_files['A1'].font = Font(name='Arial', size=14, bold=True)
        ws_files.row_dimensions[1].height = 25
        
        # Ghi dữ liệu
        files_df = pd.DataFrame(file_info)
        headers = ['Tên file', 'Số dòng', 'MY_SHOP', 'COMPETITOR']
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws_files.cell(row=2, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
        
        for row_idx, (_, row_data) in enumerate(files_df.iterrows(), start=3):
            for col_idx, value in enumerate(row_data.values, start=1):
                cell = ws_files.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
        
        # Điều chỉnh độ rộng cột
        ws_files.column_dimensions['A'].width = 40
        for col in ['B', 'C', 'D']:
            ws_files.column_dimensions[col].width = 15
    
    # Lưu vào BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output




def _create_swot_sheet(ws, title: str, items: List[Dict], columns: List[str]):
    """
    Tạo sheet cho một nhóm SWOT
    
    Args:
        ws: Worksheet object
        title: Tiêu đề sheet
        items: List các items
        columns: Danh sách các cột cần hiển thị
    """
    # Header
    ws['A1'] = title
    ws['A1'].font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='1f77b4', end_color='1f77b4', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells(f'A1:{chr(64 + len(columns))}1')
    ws.row_dimensions[1].height = 30
    
    # Column headers
    column_headers = {
        'topic': 'Chủ đề',
        'description': 'Mô tả',
        'impact': 'Mức độ ảnh hưởng',
        'root_cause': 'Nguyên nhân gốc rễ',
        'action_idea': 'Gợi ý hành động',
        'risk_level': 'Mức độ rủi ro',
        # Enterprise columns
        'priority_score': 'Điểm ưu tiên',
        'leverage_strategy': 'Chiến lược tận dụng',
        'kpi_metrics': 'KPIs',
        'mitigation_plan': 'Kế hoạch khắc phục',
        'improvement_cost': 'Chi phí cải thiện',
        'market_size': 'Quy mô thị trường',
        'time_to_capture': 'Thời gian nắm bắt',
        'required_investment': 'Đầu tư cần thiết',
        'probability': 'Xác suất',
        'severity': 'Mức độ nghiêm trọng',
        'contingency_plan': 'Kế hoạch ứng phó'
    }

    
    for col_idx, col in enumerate(columns, start=1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = column_headers.get(col, col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # Data rows
    for row_idx, item in enumerate(items, start=3):
        for col_idx, col in enumerate(columns, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            value = item.get(col, '')
            # Convert list values to comma-separated string
            if isinstance(value, list):
                value = ', '.join(str(v) for v in value)
            elif value is None:
                value = ''
            cell.value = value
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

    
    # Điều chỉnh độ rộng cột
    ws.column_dimensions['A'].width = 25  # Chủ đề
    ws.column_dimensions['B'].width = 50  # Mô tả
    for col in ['C', 'D', 'E']:
        ws.column_dimensions[col].width = 25
    
    # Điều chỉnh chiều cao hàng
    for row in range(3, len(items) + 3):
        ws.row_dimensions[row].height = 60
